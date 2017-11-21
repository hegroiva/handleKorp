#!python
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

from openpyxl.reader.excel import load_workbook

from itertools import chain
import datetime
from openpyxl.workbook import Workbook
import os
import re

def listSentNewRids():

    lstRids = []
    path = "Z:\\Documents\\Työ\\Output\\"
    rids = getSentRidInformationFromFolder(path + "16-50", "fixed", new=True)
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "comp", new=True)
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "nolla", "_nolla_", new=True)
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "nsubj_dobj", new=True)
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "poss", new=True)
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "preps", new=True)
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "sillsallad1", new=True)
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "sillsallad2", new=True)
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "varia", new=True)
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "varia2plus", new=True)
    lstRids.extend(rids)
    
def listSentRids():

    lstRids = []
    
    path = "Z:\\Documents\\Työ\\Output\\"
    rids = getSentRidInformationFromFolder(path + "16-50", "fixed")
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "comp")
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "nolla", "_nolla_")
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "nsubj_dobj")
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "poss")
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "preps")
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "sillsallad1")
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "sillsallad2")
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "varia")
    lstRids.extend(rids)
    rids = getSentRidInformationFromFolder(path + "varia2plus")
    lstRids.extend(rids)
    #rids = getSentRidInformationFromFolder(path + "varia3")
    #lstRids.extend(rids)

    return lstRids

def getSentRidInformationFromFolder(foldername, additionalReq = ""):

    rids = []
    
    for filename in os.listdir(foldername):
        if additionalReq in filename:
            
            f = open(foldername + "\\" + filename, "r", encoding="cp1252", newline='')
            text = f.read()
            f.close()

            rid = None
            sentences = 0
            lines = text.split("\r\n")
            for line in lines:
                if line.startswith("rid="):
                    # vanha rid
                    if rid is not None:
                        rids.append((rid, sentences, filename))
                    rid = re.search("[0-9]+", line).group()
                    sentences = 0
                elif "KLK_FI_" in line:
                    sentences += 1

            rids.append((rid, sentences, filename))

    return rids

def listAnnotations():

    dictTranslations = getDictTranslations()
    
    lstRids = getAnnotatedRidInformationByFolders(containsRids=True)

    rids = getAnnotatedRidInformationByFolders(folder = "Z:\\Documents\\Työ\\Annotated\\5-15", \
                                      containsRids = False, \
                                      dictTranslations=dictTranslations)

    lstRids.extend(rids)

    return lstRids


def getAnnotatedRidInformationByFolders(folder = "Z:\\Documents\\Työ\\Annotated", containsRids = True, dictTranslations = {}):

    lstRids = []
        
    if os.path.isdir(folder):
        for filename in os.listdir (folder):
            #if (("zips" in filename) == False) and \
            #   ((os.path.isdir(folder + "\\" + filename)) == False) and \
            #   ("varia" in filename):
            if (("zips" in filename) == False) and \
               ((os.path.isdir(folder + "\\" + filename)) == False):
                sheets = getFirstSheet(folder + "\\" + filename)
                rids = getAnnotatedRidInformation(sheets, containsRids, dictTranslations=dictTranslations)
                lstRids.extend(rids)
                
    else:
        sheets = getFirstSheet(folder + "\\" + filename)
        lstRids = getAnnotatedRidInformation(sheets, containsRids, dictTranslations=dictTranslations)
                       

    return lstRids

def getSomeColumns(sheet, limit):

    if sheet.get_highest_column() <= limit:
        return sheet.columns
    else:
        cols = []
        for col_idx in range(limit):
            cells = sheet.get_squared_range(col_idx + 1, 1, col_idx + 1, sheet.max_row)
            col = chain.from_iterable(cells)
            cols.append(tuple(col))
        return tuple(cols)
    
    
def getAnnotatedRidInformation(sheets, containsRids = True, dictTranslations = {}):
    """
    Returns a tuple of rid-specific information:
    rid, frame, ENLU, left_context, FEE, right_context, \
    FILU, filename, sent, annotated, metaphors, counter_examples, disqualified
    """

    lstRids = []
    dictFrames = {}
    dictFrameFILUs = {}

    sent = 0
    annotated = 0
    counter_examples = 0
    metaphors = 0
    disqualified = 0
 
    frame = None
    ENLU = None
    FILU = None
    left_context = ""
    FEE = ""
    right_context = ""
    words = []
    hasSentences = False

    boolAnnotated = False
    boolMetaphors = False
    boolCounter_examples = False


    FEE_cleared = False
    
    if type(sheets) != list:
        sheets = [sheets]

    for sheet, filename in sheets:
        print(filename)

        try:
            max_row = sheet.get_highest_row()
            max_col = sheet.get_highest_column()
            
            #cols = sheet.columns
            cols = getSomeColumns(sheet, 15)
            
            rid = None
            # Ensimmäisessä kolumnissa pitäisi olla olennainen jäsennetty tieto
            for cell in cols[0]:
                # Varotoimenpide, ettei tarvitse koko ajan tarkistaa puuttuvaa arvoa
                if (cell.value == None) or (cell.value == ""):
                    continue

                # Excel muuttaa itsekseen merkkijonoja luvuiksi tai päivämääriksi
                # Tässäkin aineistossa ongelmaa on esiintynyt
                if type(cell.value) != str:
                    cellValue = str(cell.value)
                else:
                    cellValue = cell.value
                    
                # En tiedä miksi tuosta täytyy vähentää yksi
                # Merkillinen API, jos minulta kysytään
                row = (cell.row -1)
                
                # Rid on ilmaistu tiedostossa (normaalitapaus)
                if containsRids:

                    if (re.match("[0-9]+frame=", cellValue)) or \
                       (re.match("rid=[0-9]+", cellValue)):
                        
                        if (frame != None):

                            if sent != 0 and annotated == 0:
                                disqualified = sent
                                
                            lstRids.append((rid, \
                                           frame, \
                                           ENLU, \
                                           left_context, \
                                           FEE, \
                                           right_context, \
                                           FILU, \
                                           filename, \
                                           sent, \
                                           annotated, \
                                           metaphors, \
                                           counter_examples, \
                                           disqualified))
                            if filename == "" and annotated != 0:
                                print(lstRids[-1])
                                raise Exception
                            elif sent == 0 and annotated != 0:
                                print(lstRids[-1])
                                raise Exception
                            elif right_context == "kummankin naisen":
                                print(lstRids[-1])
                                raise Exception
                            elif sent < annotated:
                                print(lstRids[-1])
                                raise Exception
                            
                            if (not rid is None) and \
                               not re.match("^[0-9]+$", rid):
                                print("Rid failure")
                                print(frame + "_" + lexunit)
                                raise Exception

                            
                            frame = ""
                            ENLU = ""
                            left_context = ""
                            FEE = ""
                            right_context = ""
                            FILU = ""
                            #filename = ""
                            sent = 0
                            annotated = 0
                            metaphors = 0
                            counter_examples = 0
                            disqualified = 0
                            frame_total = 0
                            frame_FILU_total = 0

                            boolAnnotated = False
                            boolMetaphors = False
                            boolCounter_examples = False

                            FEE_cleared = False

                        rid = getValue(cellValue, "rid")
                        
                        if "frame" in cellValue:
                            frame = getValue(cellValue, "frame")
                        
                            
                # Rid puuttuu tiedostoista (vain varhaisimmat tiedostot)
                else:

                    if (cellValue.startswith('frame="')):
                        if (frame is not None):

                            if sent != 0 and annotated == 0:
                                disqualified = sent

                            if sent < annotated:
                                print(lstRids[-1])
                                raise Exception

                            lstRids.append((rid, \
                                           frame, \
                                           ENLU, \
                                           left_context, \
                                           FEE, \
                                           right_context, \
                                           FILU, \
                                           filename, \
                                           sent, \
                                           annotated, \
                                           metaphors, \
                                           counter_examples, \
                                           disqualified))
                            
                            if not re.match("^[0-9]+$", rid):
                                print("Rid failure")
                                print(frame + "_" + lexunit)
                                raise Exception

                            rid = getValue(cellValue, "rid")
                            frame = ""
                            ENLU = ""
                            left_context = ""
                            FEE = ""
                            right_context = ""
                            FILU = ""
                            #filename = ""
                            sent = 0
                            annotated = 0
                            metaphors = 0
                            counter_examples = 0
                            disqualified = 0
                            frame_total = 0
                            frame_FILU_total = 0

                            boolAnnotated = False
                            boolMetaphors = False
                            boolCounter_examples = False

                            FEE_cleared = False
                        frame = getValue(cellValue, "frame")
                        rid = ""

                
                # Esimerkin Frame evoking element
                if (cell.value in ("FEE", "<FEE>")) or \
                   ("<FEE>" in cellValue):
                    if boolAnnotated == False:
                        annotated += 1
                    boolAnnotated = True
                elif (cell.value in ("FEEM", "<FEEM>")):
                    if boolMetaphors == False:
                        metaphors += 1
                    boolMetaphors = True
                # Uuden Korpista bongatun lauseen otsikko
                elif ("KLK_FI_" in cellValue):
                    boolAnnotated = False
                    boolMetaphors = False
                    boolCounter_examples = False
                    sent += 1
                # Vastaesimerkki
                elif (cellValue in ("VE", "<VE>")):
                    if boolCounter_examples == False:
                        counter_examples += 1
                    boolCounter_examples == True
                # Solu kertoo framen nimen
                elif (cellValue.startswith('frame="')):
                    frame = getValue(cellValue, "frame")
                # Solu kertoo lexical unitin nimen
                elif ("lexunit_name" in cellValue):
                    ENLU = getValue(cellValue, "lexunit_name")
                # fes = elementit: monta solua alkaen vasemmalta
                elif (cellValue.startswith('fes="')):
                    vals = listNeighboringCells(sheet, row, 1)
                    fes = getValue(vals, "fes")
                # fes_rel = elementit + dependenssit
                elif (cellValue.startswith('fes_rel="')):
                    vals = listNeighboringCells(sheet, row, 1)
                    fes_rel = getValue(vals, "fes_rel")
                # fes_pos_fn = elementit + dependenssit + pos
                elif (cellValue.startswith('fes_pos_fn="')):
                    vals = listNeighboringCells(sheet, row, 1)
                    fes_pos_fn = getValue(vals, "fes_pos_fn")
                # Solussa kerrotaan kehyselementti tai useampi
                elif ((cellValue.startswith("<") or cellValue.endswith(">"))) and \
                     ((cellValue.lower().startswith("<sentence "))== False) :
                    frameElements = getElementList(cellValue)
                    if "VE" in frameElements:
                        if boolCounter_examples == False:
                            counter_examples += 1
                        boolCounter_examples = True
                    if "FEEM" in frameElements:
                        if boolMetaphors == False:
                            metaphors += 1
                        boolMetaphors = True
                    if "FEE" in frameElements:
                        if boolAnnotated == False:
                            annotated += 1
                        boolAnnotated = True
                elif (cellValue.startswith('elemtype="')):

                    # Selvitetään vasemman ja oikean kontekstin rivimäärät
                    if (FEE_cleared == False):
                        lt_context = 0
                        FEE_found = False
                        rt_context = 0
                        i = 0
                        while True:
                            if getCellValue(sheet, (row+i), 1) is not None:
                                if getCellValue(sheet, (row+i), 1).startswith('elemtype="fee'):
                                    feeFound = True
                                else:
                                    if FEE_found:
                                        rt_context += 1
                                    else:
                                        lt_context += 1
                            i += 1
                            break
                            

                        FEE_cleared = True
                        
                    # NYT TARKKANA!
                    # Viimeisen elemtype-rivin jälkeen
                    # yritetään metsästää suomenkielisiä käännöksiä
                    
                    # Kaksi vasemmanpuoleista kolumnia on tyhjiä,
                    # Kolmannessa englanninkielinen elementti,
                    # Neljännessä suomenkielinen käännös
                    # Tällaisia kokonaisuuksia voi olla allekkain paljonkin
                    row = (cell.row -1)
                    if getCellValue(sheet, row + 1, 3) == None and \
                       getCellValue(sheet, row + 2, 3) == None and \
                       getCellValue(sheet, row + 1, 1) == None and \
                       getCellValue(sheet, row + 2, 1) == None and \
                       getCellValue(sheet, row + 1, 2) == None:

                        left_context = listNeighboringCellsNtimes(sheet, row + 2, 4, lt_context, \
                                                                  direction = "down", \
                                                                  separator = "\t")
                        FEE = listNeighboringCellsNtimes(sheet, (row + 2 + lt_context), 4, lt_context, \
                                                         direction = "down", \
                                                         separator = "\t")
                        right_context = listNeighboringCellsNtimes(sheet, (row + 4 + lt_context), 3, rt_context, \
                                                                   direction = "down", \
                                                                   separator = "\t")
                        #lstWords = words.split("\t")
                       
                       
                        
                    # Nämä tiedot edeltävät elemtypeä, joten ne ovat jo kasassa
                    # Ne saattavat auttaa debuggauksessa, mutta lienevät turhia
                    frameText = frame + "::" + \
                                ENLU + "::" + \
                                fes + "::" + \
                                fes_rel + "::" + \
                                fes_pos_fn


                    # Suurin osa tiedostoista sisältää rid:n mutta eivät aivan kaikki
                    # Loput täytyy selvittää elementti- ja F+LU -tiedoilla
                    # Ongelmaksi saattaa tulla se, että elementtejä on vaihdettu
                    if containsRids == False:
                        rid = dictTranslations[frameText]

            # Koska append tapahtuu vasta seuraavan rid:n löytyessä,
            # täytyy muistaa lisätä myös iteroinnin päätteeksi
            lstRids.append((rid, \
                           frame, \
                           ENLU, \
                           left_context, \
                           FEE, \
                           right_context, \
                           FILU, \
                           filename, \
                           sent, \
                           annotated, \
                           metaphors, \
                           counter_examples, \
                           disqualified))
                        
        except Exception as exc:
            print(filename)
            print(row)
            print(frameText)
            raise Exception(str(exc))
            
                                                                        

    return lstRids


def createWorkBookExample():
    wb = Workbook()

    dest_filename = r'empty_book.xlsx'

    ws = wb.worksheets[0]

    ws.title = "range names"
    for col_idx in range(1, 40):
        col = get_column_letter(col_idx)
        for row in range(1, 600):
            ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

    ws = wb.create_sheet()

    ws.title = 'Pi'

    ws.cell('F5').value = 3.14

    wb.save(filename = dest_filename)


def loadWorkbooks(folder = "Z:\\Documents\\Työ\\Annotated"):

    sheets = []
    if os.path.isdir(folder):
        for filename in os.listdir (folder):
            if (("zips" in filename) == False) and \
               ((os.path.isdir(folder + "\\" + filename)) == False):
            
                wb = load_workbook(folder + "\\" + filename)
                sheet_ranges = wb.worksheets[0]
                sheets.append([sheet_ranges, filename])
            else:
                continue
    else:
        wb = load_workbook(folder)
        sheet_ranges = wb.worksheets[0]
        sheets.append([sheet_ranges, folder])
        
    return sheets

def getFirstSheet(filename):

    sheets = []
    wb = load_workbook(filename)
    sheet_ranges = wb.worksheets[0]
    sheets.append([sheet_ranges, filename])

    return sheets



def countAnnotations():

    dictTranslations = getDictTranslations()
    
    dictFrameLexunits, dictCounterExamples, \
                       dictMetaphors, \
                       dictFrameElements, \
                       lstUnsentRids, \
                       lstAnnotatedFrames, \
                       lstAnnotatedFrameLexunits, \
                       lstAnnotatedRids, \
                       = countFolders(containsRids = True)

    dictFrameLexunits2, dictCounterExamples2, \
                        dictMetaphors2, \
                        dictFrameElements2, \
                        lstUnsentRids2, \
                        lstAnnotatedFrames2, \
                        lstAnnotatedFrameLexunits2, \
                        lstAnnotatedRids2, \
                        = countFolders(folder = "Z:\\Documents\\Työ\\Annotated\\5-15", \
                                       containsRids = False, \
                                       dictTranslations=dictTranslations)

    dictFrameLexunits.update(dictFrameLexunits2)
    dictCounterExamples.update(dictCounterExamples2)
    dictMetaphors.update(dictMetaphors2)
    dictFrameElements.update(dictFrameElements2)
    lstUnsentRids += lstUnsentRids2
    lstAnnotatedFrames += lstAnnotatedFrames2
    lstAnnotatedFrameLexunits += lstAnnotatedFrameLexunits2
    lstAnnotatedRids += lstAnnotatedRids2

    return dictFrameLexunits, \
           dictCounterExamples, \
           dictMetaphors, \
           dictFrameElements, \
           lstUnsentRids, \
           lstAnnotatedFrames, \
           lstAnnotatedFrameLexunits, \
           lstAnnotatedRids
    
    
def countFolders(folder = "Z:\\Documents\\Työ\\Annotated", containsRids = True, dictTranslations = {}):

    dictFrameLexunits = {}
    dictCounterExamples = {}
    dictMetaphors = {}
    dictFrameElements = {}
    lstUnsentRids = []
    lstAnnotatedFrames = []
    lstAnnotatedFrameLexunits = []
    lstAnnotatedRids = []
    
    
    if os.path.isdir(folder):
        for filename in os.listdir (folder):
            if (("zips" in filename) == False) and \
               ((os.path.isdir(folder + "\\" + filename)) == False):

                sheets = getFirstSheet(folder + "\\" + filename)
                dictFrameLexunits2, dictCounterExamples2, \
                       dictMetaphors2, \
                       dictFrameElements2, \
                       lstUnsentRids2, \
                       lstAnnotatedFrames2, \
                       lstAnnotatedFrameLexunits2, \
                       lstAnnotatedRids2, \
                       = countFrameLexunitSentences(sheets, containsRids, dictTranslations=dictTranslations)

                dictFrameLexunits.update(dictFrameLexunits2)
                dictCounterExamples.update(dictCounterExamples2)
                dictMetaphors.update(dictMetaphors2)
                dictFrameElements.update(dictFrameElements2)
                lstUnsentRids += lstUnsentRids2
                lstAnnotatedFrames += lstAnnotatedFrames2
                lstAnnotatedFrameLexunits += lstAnnotatedFrameLexunits2
                lstAnnotatedRids += lstAnnotatedRids2 
                
    else:
        sheets = getFirstSheet(folder + "\\" + filename)
        dictFrameLexunits, dictCounterExamples, \
                       dictMetaphors, \
                       dictFrameElements, \
                       lstUnsentRids, \
                       lstAnnotatedFrames, \
                       lstAnnotatedFrameLexunits, \
                       lstAnnotatedRids, \
                       = countFrameLexunitSentences(sheets, containsRids, dictTranslations=dictTranslations)

    return dictFrameLexunits, \
           dictCounterExamples, \
           dictMetaphors, \
           dictFrameElements, \
           lstUnsentRids, \
           lstAnnotatedFrames, \
           lstAnnotatedFrameLexunits, \
           lstAnnotatedRids
    
def countFrameLexunitSentences(sheets, containsRids = True, dictTranslations = {}):
    """
    Returns a tuple of four dictionaries and four lists:
      1st: dictFrameLexunits[frame + "_" + lexunit] = # of annotated sentences
      2nd: dictCounterExamples[frame + "_" + lexunits] = # of counter examples
      3rd: dictMetaphors[frame + "_" + lexunits] = # of metaphoric FEEs
      4th: dictFrameElements[frame] = collected frame element names
      5th: lstUnsentRids
      6th: lstAnnotatedFrames
      7th: lstAnnotatedFrameLexunits
      8th: lstAnnotatedRids
    """

    #dictTranslations = getDictTranslations()

    dictMetaphors = {}
    dictCounterExamples = {}
    dictFrameLexunits = {}
    dictFrameElements = {}
    lstUnsentRids = []
    lstAnnotatedFrames = []
    lstAnnotatedFrameLexunits = []
    lstAnnotatedRids = []

    examples = 0
    counterExamples = 0
    metaphors = 0
    frameElements = []
    frame = None
    lexunit = None
    words = []
    hasSentences = False
    
    if type(sheets) != list:
        sheets = [sheets]

    for sheet, filename in sheets:
        print(filename)

        try:
            max_row = sheet.get_highest_row()
            max_col = sheet.get_highest_column()
            
            cols = sheet.columns
            foundFEE = False
            rid = None
            # Ensimmäisessä kolumnissa pitäisi olla olennainen jäsennetty tieto
            for cell in cols[0]:
                # Varotoimenpide, ettei tarvitse koko ajan tarkistaa puuttuvaa arvoa
                if (cell.value == None) or (cell.value == ""):
                    continue

                # Excel muuttaa itsekseen merkkijonoja luvuiksi tai päivämääriksi
                # Tässäkin aineistossa ongelmaa on esiintynyt
                if type(cell.value) != str:
                    cellValue = str(cell.value)
                else:
                    cellValue = cell.value
                    
                # En tiedä miksi tuosta täytyy vähentää yksi
                # Merkillinen API, jos minulta kysytään
                row = (cell.row -1)
                
                # Rid on ilmaistu tiedostossa (normaalitapaus)
                if containsRids:
                    # Ensimmäinen solu, aloittaa esimerkin

                    # debug
                    #if filename != None:
                        #print(filename)
                    
                            

                    if (re.match("[0-9]+frame=", cellValue)) or \
                       (re.match("rid=[0-9]+", cellValue)):
                        
                        if (foundFEE == False) and \
                           (frame != None) and \
                           (hasSentences):
                            #print("Ei löytynyt FEE:tä")
                            #print(frame + "_" + lexunit)
                            #print(words)
                            #print(frameText)
                            #print(rid)
                            lstUnsentRids.append(rid)
                            #print(frame)
                            #print(rid)
                            #print(filename)
                            #raise Exception
                        # Siis 'frame' tarkoittaa vielä edellistä kehystä
                        # Koska FEE:tä ei löytynyt, ei sitä ole annotoitu
                        # siksi 'elif', eikä 'if'
                        elif ((frame in lstAnnotatedFrames) == False) and \
                           (frame != None):
                            lstAnnotatedFrames.append(frame)
                            lstAnnotatedFrameLexunits.append(frame + "_" + lexunit)
                            lstAnnotatedRids.append(rid)
                            if (not rid is None) and \
                               not re.match("^[0-9]+$", rid):
                                print("Rid failure")
                                print(frame + "_" + lexunit)
                                raise Exception
                        elif (frame != None) and (lexunit != None) and \
                             (((frame + "_" + lexunit) in lstAnnotatedFrameLexunits) == False):
                            lstAnnotatedFrameLexunits.append(frame + "_" + lexunit)
                            lstAnnotatedRids.append(rid)
                        foundFEE = False
                        hasSentences = False
                        rid = getValue(cellValue, "rid")
                        
                        if "frame" in cellValue:
                            frame = getValue(cellValue, "frame")
                        
                            
                # Rid puuttuu tiedostoista (vain varhaisimmat tiedostot)
                else:
                    # debug
                    #if filename != None:
                    #    print(filename)

                    if (cellValue.startswith('frame="')):
                        if (foundFEE == False) and \
                           (frame is not None) and \
                           (hasSentences):
                            #print("Ei löytynyt FEE:tä")
                            #print(frame + "_" + lexunit)
                            #print(words)
                            #print(frameText)
                            #print(rid)
                            lstUnsentRids.append(rid)
                            #raise Exception
                        # Siis 'frame' tarkoittaa vielä edellistä kehystä
                        # Koska FEE:tä ei löytynyt, ei sitä ole annotoitu
                        # siksi 'elif', eikä 'if'
                        elif ((frame in lstAnnotatedFrames) == False) and \
                           (frame != None):
                            lstAnnotatedFrames.append(frame)
                            lstAnnotatedFrameLexunits.append(frame + "_" + lexunit)
                            lstAnnotatedRids.append(rid)
                            if not re.match("^[0-9]+$", rid):
                                print("Rid failure")
                                print(frame + "_" + lexunit)
                                raise Exception
                        elif (frame != None) and (lexunit != None) and \
                            ((frame + "_" + lexunit in lstAnnotatedFrameLexunits) == False):
                            lstAnnotatedFrameLexunits.append(frame + "_" + lexunit)
                            lstAnnotatedRids.append(rid)
                        foundFEE = False
                        hasSentences = False
                        frame = getValue(cellValue, "frame")
                        rid = ""

                
                # Esimerkin Frame evoking element
                if (cell.value in ("FEE", "<FEE>")) or \
                   ("<FEE>" in cellValue):
                    examples += 1
                    addDictValue(dictFrameLexunits, frame + "_" + lexunit, 1)
                    foundFEE = True
                elif (cell.value in ("FEEM", "<FEEM>")):
                    metaphors += 1
                    addDictValue(dictMetaphors, frame + "_" + lexunit, 1)
                    addDictValue(dictFrameLexunits, frame + "_" + lexunit, 1)
                    foundFEE = True
                # Uuden Korpista bongatun lauseen otsikko
                elif ("KLK_FI_" in cellValue):
                    hasSentences = True
                # Vastaesimerkki
                elif (cellValue in ("VE", "<VE>")):
                    counterExamples += 1
                    addDictValue(dictCounterExamples, frame + "_" + lexunit, 1)
                    #foundFEE = True
                # Solu kertoo framen nimen
                elif (cellValue.startswith('frame="')):
                    frame = getValue(cellValue, "frame")
                # Solu kertoo lexical unitin nimen
                elif ("lexunit_name" in cellValue):
                    lexunit = getValue(cellValue, "lexunit_name")
                # fes = elementit: monta solua alkaen vasemmalta
                elif (cellValue.startswith('fes="')):
                    vals = listNeighboringCells(sheet, row, 0)
                    fes = getValue(vals, "fes")
                # fes_rel = elementit + dependenssit
                elif (cellValue.startswith('fes_rel="')):
                    vals = listNeighboringCells(sheet, row, 0)
                    fes_rel = getValue(vals, "fes_rel")
                # fes_pos_fn = elementit + dependenssit + pos
                elif (cellValue.startswith('fes_pos_fn="')):
                    vals = listNeighboringCells(sheet, row, 0)
                    fes_pos_fn = getValue(vals, "fes_pos_fn")
                # Solussa kerrotaan kehyselementti tai useampi
                elif ((cellValue.startswith("<") or cellValue.endswith(">"))) and \
                     ((cellValue.lower().startswith("<sentence "))== False) :
                    frameElements = getElementList(cellValue)
                    if "VE" in frameElements:
                        counterExamples += 1
                        addDictValue(dictCounterExamples, frame + "_" + lexunit, 1)
                    if "FEE" in frameElements:
                        examples += 1
                        addDictValue(dictFrameLexunits, frame + "_" + lexunit, 1)
                        foundFEE = True
                    addListValuesToDict(dictFrameElements, frame, frameElements)
                elif (cellValue.startswith('elemtype="')):
                    # NYT TARKKANA!
                    # Viimeisen elemtype-rivin jälkeen
                    # yritetään metsästää englanninkielisiä originaaleja
                    
                    # Kaksi vasemmanpuoleista kolumnia on tyhjiä,
                    # Kolmannessa englanninkielinen elementti,
                    # Neljännessä suomenkielinen käännös
                    # Tällaisia kokonaisuuksia voi olla allekkain paljonkin
                    row = (cell.row -1)
                    if getCellValue(sheet, row + 1, 0) == None and \
                       getCellValue(sheet, row + 2, 0) == None and \
                       getCellValue(sheet, row + 1, 1) == None and \
                       getCellValue(sheet, row + 2, 1) == None and \
                       getCellValue(sheet, row + 1, 2) == None:

                        words = listNeighboringCells(sheet, row + 2, 2, \
                                                     direction = "down", \
                                                     separator = "\t")
                        lstWords = words.split("\t")
                       

                    # Nämä tiedot edeltävät elemtypeä, joten ne ovat jo kasassa
                    # Ne saattavat auttaa debuggauksessa, mutta lienevät turhia
                    frameText = frame + "::" + \
                                lexunit + "::" + \
                                fes + "::" + \
                                fes_rel + "::" + \
                                fes_pos_fn

                    # rid:tä ei välttämättä tarvita enää lainkaan
                    # frameText sisältää kehyksen ja LU:n nimet

                    # Suurin osa tiedostoista sisältää rid:n mutta eivät aivan kaikki
                    # Loput täytyy selvittää elementti- ja F+LU -tiedoilla
                    # Ongelmaksi saattaa tulla se, että elementtejä on vaihdettu
                    if containsRids == False:
                        rid = dictTranslations[frameText]
                        
        except Exception as exc:
            print(filename)
            print(row)
            print(frameText)
            raise Exception(str(exc))
            
                                                                        

    return dictFrameLexunits, \
           dictCounterExamples, \
           dictMetaphors, \
           dictFrameElements, \
           lstUnsentRids, \
           lstAnnotatedFrames, \
           lstAnnotatedFrameLexunits, \
           lstAnnotatedRids


def getCellValue(sheet, row, column):
    """
    Otetaan yksittäisen solun arvo.
    Oikeastaan tämä funktio on lähinnä tilan säästämiseksi
    sheet = openpyxl.worksheet
    row = solun rivinumero (zero-base)
    column = solun kolumninumero (zero-base)
    """

    if sheet.cell(row = row, column = column) == None:
        return None
    
    if sheet.cell(row = row, column = column).value == None:
        return None

    # En nyt ole ihan saletti, voiko solun arvona olla tyhjä merkkijono
    if sheet.cell(row = row, column = column).value == "":
        return None

    # Muutetaan vielä varmuudeksi merkkijonoksi
    # Excelillä on taipumus muuttaa datatyyppiä usein katkerasti yllättäen
    return str(sheet.cell(row = row, column = column).value)


def listNeighboringCells(sheet, row, column, \
                         direction = "right", \
                         separator = " "):
    """
    Listataan vierekkäisten solujen arvot ja erotellaan ne soveliaalla erotinmerkillä
    sheet = openpyxl.worksheet
    row = alkusolun rivinumero (zero-base)
    column = alkusolun kolumninumero (zero-base)
    direction = "left", "right", "up", "down"; mihin suuntaan alkusolusta lähdetään
    separator = erotinmerkki(jono), jolla solujen arvot erotetaan toisistaan
    """
    
    # Alkuarvo, johon sitten lisätään muiden solujen arvoja
    try:
        ret = getCellValue(sheet, row, column)
    except:
        print(sheet)
        print(row)
        print(column)
        raise Exception

    # Rauhallinen exitus
    if ret == None:
        return ret
    
    # Riskaabelihko while:
    # on muistettava muuttaa arvoja ja lopetettava iterointi, kun arvoa ei enää ole
    while True:
        # seuraava rivi tai kolumni
        if direction == "right":
            column += 1
        elif direction == "left":
            column -= 1
        elif direction == "down":
            row += 1
        elif direction == "up":
            row -= 1
        
        cellValue = getCellValue(sheet, row, column)
        if cellValue != None:
            ret += (separator + cellValue)
        else:
            # Kun arvoa ei ole, lopetetaan iterointi
            break
                                 
    return ret            


def listNeighboringCellsNtimes(sheet, row, column, times, \
                         direction = "right", \
                         separator = " "):
    """
    Listataan vierekkäisten solujen arvot ja erotellaan ne soveliaalla erotinmerkillä
    sheet = openpyxl.worksheet
    row = alkusolun rivinumero (zero-base)
    column = alkusolun kolumninumero (zero-base)
    direction = "left", "right", "up", "down"; mihin suuntaan alkusolusta lähdetään
    separator = erotinmerkki(jono), jolla solujen arvot erotetaan toisistaan
    """
    
    # Alkuarvo, johon sitten lisätään muiden solujen arvoja
    ret = getCellValue(sheet, row, column)

    # Rauhallinen exitus
    if ret == None:
        return ret
    
    # Riskaabelihko while:
    # on muistettava muuttaa arvoja ja lopetettava iterointi, kun arvoa ei enää ole
    i = 0
    while i in range(0, times):
        # seuraava rivi tai kolumni
        if direction == "right":
            column += 1
        elif direction == "left":
            column -= 1
        elif direction == "down":
            row += 1
        elif direction == "up":
            row -= 1
        
        cellValue = getCellValue(sheet, row, column)
        if cellValue != None:
            ret += (separator + cellValue)
        else:
            # Kun arvoa ei ole, lopetetaan iterointi
            break

        i += 1
         
    return ret                        
    
    
 
def getDictTranslations(origPath):
    allTranslations = []
    for j in range(1,103):
        jStr = ("000" + str(j))[-3:]
        transFile = open (origPath + "\\" + \
                        "frames_" + jStr + ".xml", \
                        "r", encoding="utf8")
        translationText = transFile.read()
        translations = translationText.split("-------------------- END OF SENTENCE --------------------")
        for translation in translations:
            if ("<sentence" in translation):
                allTranslations.append(translation)
    
    dictTranslations = getTranslationDict(allTranslations)
    return dictTranslations
    

def getTranslationDict(translations):
    retDict = {}
    for translation in translations:
        frameText = getValue(translation, "frame")
        lexunit_name = getValue(translation, "lexunit_name")
        fes = getValue(translation, "fes")
        fes_rel = getValue(translation, "fes_rel")
        fes_pos_fn = getValue(translation, "fes_pos_fn")
        rid = getValue(translation, "rid")

        # Hieman nurinkurista, eikö vain?
        retDict[frameText + "::" + \
                lexunit_name + "::" + \
                fes +"::" + \
                fes_rel + "::" + \
                fes_pos_fn] = rid

    return retDict

def getElementList(string):

    # Poistetaan ensin turhat merkit
    # Tässä oli ensin hienovaraisempi lähestymistapa,
    # mutta erottimien moninaisuus sekoitti suunnitelmat

    separatorStrings = ">/FEE", "<", ">", "?", "/", "|", " "
    for separatorString in separatorStrings:
        string = string.replace(separatorString, "===")
    
    # Samaan sanaan voi sisältyä useita elementtejä
    stringParts = string.split("===")
    elements = []

    for stringPart in stringParts:
        if ((stringPart in elements) == False):

            if stringPart != "":
                elements.append(stringPart)

    return elements
        
def addListValuesToDict(dictionary, key, valuesToAdd):
    """
    Add values to dictionary[key], if they're missing
    No return value
    valuesToAdd = a list of values
    """

    if key in dictionary.keys():
        lstVals = dictionary.pop(key)
    else:
        lstVals = []
        
    for valueToAdd in valuesToAdd:
        if ((valueToAdd in lstVals) == False):
            if valueToAdd.upper() not in ("FEE", "VE"):
                lstVals.append(valueToAdd)
    dictionary[key] = lstVals

            
def addDictValue(dictionary, key, valueToAdd):

    if key in dictionary.keys():
        val = dictionary.pop(key)
        val += valueToAdd
        dictionary[key] = val
    else:
        dictionary[key] = valueToAdd

    return dictionary


def getValue(sentence, attributeName):
    start = "(\W|^)"
    if (attributeName == "frame") and (re.match("[0-9]+frame=\"", sentence)):
        start = ""

    if (attributeName == "frame") and (re.search('frame=".*?"', sentence)):
        return re.search('frame="(.*?)"', sentence).groups()[0]
    
    if (attributeName ==  "rid") and (re.match("[0-9]+frame=\"", sentence)):
        return re.match("[0-9]+", sentence).group()

    if (attributeName ==  "rid") and (re.match("[0-9]+\*\*\* Tästä puuttuu metatietoja", sentence)):
        return re.match("[0-9]+", sentence).group()

    if (attributeName ==  "rid") and (re.match("rid=[0-9]+", sentence)):
        return re.search("[0-9]+", sentence).group()

    if (attributeName == "rid") and (re.search("rid=[0-9]+", sentence)):
        return re.search("rid=([0-9]+)", sentence).groups()[0]
                             
    if attributeName == "korp_id":
        yearMatch = re.search("KLK_FI_[0-9]+", sentence).group()
        year = re.search("[0-9]+", yearMatch).group()
        runningIdMatch = re.search(" sentence [0-9]+", sentence).group()
        runningId = re.search("[0-9]+", runningIdMatch).group()
        #positionMatch = re.search("position [0-9]+", sentence).group()
        #positionId = re.search("[0-9]+", positionMatch).group()
        return year + "_" + runningId #+ "_" + positionMatch

    if attributeName == "fes_rel":
        match = re.search(start + attributeName + "\=\"[^\"]+", sentence)
        if match == None:
            return ""

        match = re.search("(" + attributeName + "\=\")([^\"]+)", match.group())
        if match == None:
            return ""

        
    match = re.search(start + attributeName + "\=\"[^\"]+", sentence)
    if match == None:
        return ""

    match = re.search("(" + attributeName + "\=\")([^\"]+)", match.group())
    if match == None:
        return ""

    return match.groups()[1]


def readExample1():
    wb = load_workbook(filename = r'empty_book.xlsx')

    sheet_ranges = wb.get_sheet_by_name(name = 'range names')

    print (sheet_ranges.cell('D18').value )# D18


def numberFormatsExample():
    #=== Number Formats ===
    wb = Workbook()
    ws = wb.worksheets[0]

    # set date using a Python datetime
    ws.cell('A1').value = datetime.datetime(2010, 7, 21)

    print (ws.cell('A1').style.number_format.format_code) # returns 'yyyy-mm-dd'

    # set percentage using a string followed by the percent sign
    ws.cell('B1').value = '3.14%'

    print( ws.cell('B1').value) # returns 0.031400000000000004

    print (ws.cell('B1').style.number_format.format_code) # returns '0%'


def listSearches(origPath, annotatedFolder):

    dictTranslations = getDictTranslations(origPath)
    
    lstGoodSearches, lstBadSearches = getKorpSentenceIDByFolders(folder=annotatedFolder, \
                                                                 containsRids=True, \
                                                                 dictTranslations=dictTranslations)
    #goodSearches, badSearches = getKorpSentenceIDByFolders(folder = "Z:\\Documents\\Työ\\Annotated\\5-15", \
    #                                  containsRids = False, \
    #                                  dictTranslations=dictTranslations)

    #lstGoodSearches.extend(goodSearches)
    #lstBadSearches.extend(badSearches)

    return lstGoodSearches, lstBadSearches


def getKorpSentenceIDByFolders(folder = "Z:\\Documents\\Työ\\Annotated", containsRids = True, dictTranslations = {}):

    lstGoodSearches = []
    lstBadSearches = []

            
    if os.path.isdir(folder):
        for filename in os.listdir (folder):
            if "5-15" in filename:
                serial = int(re.match("output_5-15_([0-9]+)\.", filename).groups()[0])
                if serial <= 4:
                    containsRids = False
            elif "utf8_0" in filename:
                containsRids = False

            #if (("zips" in filename) == False) and \
            #   ((os.path.isdir(folder + "\\" + filename)) == False) and \
            #   ("varia" in filename):
            if (("zips" in filename) == False) and \
               ((os.path.isdir(folder + "\\" + filename)) == False):
                sheets = getFirstSheet(folder + "\\" + filename)
                print("\t" + filename)
                goodSearches, badSearches = getKorpSentenceID(sheets, containsRids, dictTranslations=dictTranslations)
                lstGoodSearches.extend(goodSearches)
                lstBadSearches.extend(badSearches)
                
    else:
        try:
            sheets = getFirstSheet(folder)
            if "5-15" in folder:
                serial = int(re.search("output_5-15_([0-9]+)\.", folder).groups()[0])
                #print(len(sheets))
                #print(type(sheets))
                if serial <= 4:
                    containsRids = False
            elif "utf8_0" in folder:
                containsRids = False
                
            lstGoodSearches, lstBadSearches = getKorpSentenceID(sheets, containsRids, dictTranslations=dictTranslations)
        except:
           raise Exception
    return lstGoodSearches, lstBadSearches


def getKorpSentenceID(sheets, containsRids = True, dictTranslations = {}):
    """
    Returns a tuple of two lists, both containing a tuple of (rid, korp_id):
    lstGoodSearches, lstBadSearches
    """

    lstRids = []
    dictFrames = {}
    dictFrameFILUs = {}

    sent = 0
    annotated = 0
    counter_examples = 0
    metaphors = 0
    disqualified = 0
 
    frame = None
    ENLU = None
    FILU = None
    left_context = ""
    FEE = ""
    right_context = ""
    words = []
    hasSentences = False
    korp_id = ""
    lstGoodSearches = []
    lstBadSearches = []
    korp_ids = []
    
    boolAnnotated = False
    boolMetaphors = False
    boolCounter_examples = False


    FEE_cleared = False
    
    if type(sheets) != list:
        sheets = [sheets]

    for sheet, filename in sheets:
        print(filename)

        try:
            max_row = sheet.get_highest_row()
            max_col = sheet.get_highest_column()
            
            cols = getSomeColumns(sheet, 15)
            
            rid = None
            # Ensimmäisessä kolumnissa pitäisi olla olennainen jäsennetty tieto
            for cell in cols[0]:
                # Varotoimenpide, ettei tarvitse koko ajan tarkistaa puuttuvaa arvoa
                if (cell.value == None) or (cell.value == ""):
                    continue

                # Excel muuttaa itsekseen merkkijonoja luvuiksi tai päivämääriksi
                # Tässäkin aineistossa ongelmaa on esiintynyt
                if type(cell.value) != str:
                    cellValue = str(cell.value)
                else:
                    cellValue = cell.value
                    
                # En tiedä miksi tuosta täytyy vähentää yksi
                # Merkillinen API, jos minulta kysytään
                row = (cell.row -1)
                
                # Rid on ilmaistu tiedostossa (normaalitapaus)
                if containsRids:

                    if (re.match("[0-9]+frame=", cellValue)) or \
                       (re.search("rid=[0-9]+", cellValue)) or \
                       (re.match("[0-9]+\*\*\* Tästä puuttuu metatietoja \*\*\*", cellValue)):
                        
                        if (frame != None) and (frame != ""):

                            if sent != 0 and annotated == 0:
                                disqualified = sent

                            if boolAnnotated:
                                lstGoodSearches.append((rid, korp_id))

                            if annotated == 0:
                                for korp_id in korp_ids:
                                    lstBadSearches.append((rid, korp_id))
                                                      
                            if filename == "" and annotated != 0:
                                print(lstRids[-1])
                                raise Exception
                            elif sent == 0 and annotated != 0:
                                print(lstRids[-1])
                                raise Exception
                            elif sent < annotated:
                                print(lstRids[-1])
                                raise Exception
                            
                            if (not rid is None) and \
                               not re.match("^[0-9]+$", rid):
                                print("Rid failure")
                                print(frame + "_" + ENLU)
                                print("previous rid: " + str(rid))
                                print(filename)
                                raise Exception

                            
                            frame = ""
                            ENLU = ""
                            left_context = ""
                            FEE = ""
                            right_context = ""
                            FILU = ""
                            #filename = ""
                            sent = 0
                            annotated = 0
                            metaphors = 0
                            counter_examples = 0
                            disqualified = 0
                            frame_total = 0
                            frame_FILU_total = 0
                            korp_id = ""
                            korp_ids = []
                            
                            boolAnnotated = False
                            boolMetaphors = False
                            boolCounter_examples = False

                            FEE_cleared = False

                        rid = getValue(cellValue, "rid")
                        #if rid == "78392":
                        #    print("78392!!")
                        #    print(frame)
                        #    print("* * * * * *")
                        #    raise Exception
                        if "frame" in cellValue:
                            frame = getValue(cellValue, "frame")
                        elif "Tästä puuttuu metatieto" in cellValue:
                            frame = "Ei tiedossa"
                            #print("HEP")
                            #print(rid)
                            #print(frame)
                            
                # Rid puuttuu tiedostoista (vain varhaisimmat tiedostot)
                else:

                    if (cellValue.startswith('frame="')):
                        if (frame is not None):

                            if sent != 0 and annotated == 0:
                                disqualified = sent

                            if sent < annotated:
                                print(lstRids[-1])
                                raise Exception

                            if boolAnnotated:
                                lstGoodSearches.append((rid, korp_id))

                            if annotated == 0:
                                for korp_id in korp_ids:
                                    lstBadSearches.append((rid, korp_id))

                            if not re.match("^[0-9]+$", rid):
                                print("Rid failure")
                                print(frame + "_" + lexunit)
                                raise Exception

                            rid = getValue(cellValue, "rid")
                            frame = ""
                            ENLU = ""
                            left_context = ""
                            FEE = ""
                            right_context = ""
                            FILU = ""
                            #filename = ""
                            sent = 0
                            annotated = 0
                            metaphors = 0
                            counter_examples = 0
                            disqualified = 0
                            frame_total = 0
                            frame_FILU_total = 0
                            korp_id = ""
                            korp_ids = []

                            boolAnnotated = False
                            boolMetaphors = False
                            boolCounter_examples = False

                            FEE_cleared = False
                        frame = getValue(cellValue, "frame")
                        rid = ""

                
                # Esimerkin Frame evoking element
                if (cell.value in ("FEE", "<FEE>")) or \
                   ("<FEE>" in cellValue):
                    if boolAnnotated == False:
                        annotated += 1
                    boolAnnotated = True
                elif ("Tästä puuttuu metatieto" in cellValue):
                    frame = "Ei tiedossa"
                    
                elif (cell.value in ("FEEM", "<FEEM>")):
                    if boolMetaphors == False:
                        metaphors += 1
                    boolMetaphors = True
                # Uuden Korpista bongatun lauseen otsikko
                elif ("KLK_FI_" in cellValue):
                    # Lisätään edellinen korp_id hyväksi hauksi
                    if boolAnnotated:
                        lstGoodSearches.append((rid, korp_id))
                    #if rid == "78392":
                    #    print("* * * * * *")
                    #    print("78392")
                    #    print(cellValue)
                    #    raise Exception
                    # Otetaan uusi korp_id, jos se vaikka kuuluisikin huonoon hakuun
                    korp_id_to_parse = cellValue

                    # Joissain tiedostoissa arvo on jakautunut kahteen soluun
                    #row = (cell.row -1)
                    #if getCellValue(sheet, row, 1) != None:
                    #    korp_id_to_parse += ("," + getCellValue(sheet, row, 1))

                    korp_id = getValue(korp_id_to_parse, "korp_id")
                    korp_ids.append(korp_id)
                    
                    boolAnnotated = False
                    boolMetaphors = False
                    boolCounter_examples = False
                    sent += 1
                # Vastaesimerkki
                elif (cellValue in ("VE", "<VE>")):
                    if boolCounter_examples == False:
                        counter_examples += 1
                    boolCounter_examples == True
                # Solu kertoo framen nimen
                elif (cellValue.startswith('frame=')):
                    frame = getValue(cellValue, "frame")
                    #if frame == '':
                    #    print(cellValue)
                    #elif frame == 'Personal_relationship':
                    #    print("PERSONAL RELATIONSHIP")
                # Solu kertoo lexical unitin nimen
                elif ("lexunit_name" in cellValue):
                    ENLU = getValue(cellValue, "lexunit_name")
                # fes = elementit: monta solua alkaen vasemmalta
                elif (cellValue.startswith('fes="')):
                    vals = listNeighboringCells(sheet, row+1, 1)
                    fes = getValue(vals, "fes")
                # fes_rel = elementit + dependenssit
                elif (cellValue.startswith('fes_rel="')):
                    vals = listNeighboringCells(sheet, row+1, 1)
                    fes_rel = getValue(vals, "fes_rel")
                # fes_pos_fn = elementit + dependenssit + pos
                elif (cellValue.startswith('fes_pos_fn="')):
                    vals = listNeighboringCells(sheet, row+1, 1)
                    fes_pos_fn = getValue(vals, "fes_pos_fn")
                # Solussa kerrotaan kehyselementti tai useampi
                elif ((cellValue.startswith("<") or cellValue.endswith(">"))) and \
                     ((cellValue.lower().startswith("<sentence "))== False) :
                    frameElements = getElementList(cellValue)
                    if "VE" in frameElements:
                        if boolCounter_examples == False:
                            counter_examples += 1
                        boolCounter_examples = True
                    if "FEEM" in frameElements:
                        if boolMetaphors == False:
                            metaphors += 1
                        boolMetaphors = True
                    if "FEE" in frameElements:
                        if boolAnnotated == False:
                            annotated += 1
                        boolAnnotated = True
                elif (cellValue.startswith('elemtype="')):

                    # Selvitetään vasemman ja oikean kontekstin rivimäärät
                    if (FEE_cleared == False):
                        lt_context = 0
                        FEE_found = False
                        rt_context = 0
                        i = 1
                        while True:
                            if getCellValue(sheet, (row+i), 1) is not None:
                                if getCellValue(sheet, (row+i), 1).startswith('elemtype="fee'):
                                    feeFound = True
                                else:
                                    if FEE_found:
                                        rt_context += 1
                                    else:
                                        lt_context += 1
                            i += 1
                            break
                            

                        FEE_cleared = True
                        
                    # NYT TARKKANA!
                    # Viimeisen elemtype-rivin jälkeen
                    # yritetään metsästää suomenkielisiä käännöksiä
                    
                    # Kaksi vasemmanpuoleista kolumnia on tyhjiä,
                    # Kolmannessa englanninkielinen elementti,
                    # Neljännessä suomenkielinen käännös
                    # Tällaisia kokonaisuuksia voi olla allekkain paljonkin
                    row = (cell.row -1)
                    if getCellValue(sheet, row + 1, 1) == None and \
                       getCellValue(sheet, row + 2, 1) == None and \
                       getCellValue(sheet, row + 1, 2) == None and \
                       getCellValue(sheet, row + 2, 2) == None and \
                       getCellValue(sheet, row + 1, 3) == None:

                        left_context = listNeighboringCellsNtimes(sheet, row + 2, 3, lt_context, \
                                                                  direction = "down", \
                                                                  separator = "\t")
                        FEE = listNeighboringCellsNtimes(sheet, (row + 2 + lt_context), 3, lt_context, \
                                                         direction = "down", \
                                                         separator = "\t")
                        right_context = listNeighboringCellsNtimes(sheet, (row + 3 + lt_context), 3, rt_context, \
                                                                   direction = "down", \
                                                                   separator = "\t")
                        #lstWords = words.split("\t")
                       
                       
                        
                    # Nämä tiedot edeltävät elemtypeä, joten ne ovat jo kasassa
                    # Ne saattavat auttaa debuggauksessa, mutta lienevät turhia
                    #if rid == None:
                    if containsRids == False:
                        try:
                            frameText = frame + "::" + \
                                    ENLU + "::" + \
                                    fes + "::" + \
                                    fes_rel + "::" + \
                                    fes_pos_fn
                        except:
                            print(frame)
                            print(ENLU)
                            print(fes)
                            print(fes_rel)
                            print(fes_pos_fn)
                            raise Exception

                    # Suurin osa tiedostoista sisältää rid:n mutta eivät aivan kaikki
                    # Loput täytyy selvittää elementti- ja F+LU -tiedoilla
                    # Ongelmaksi saattaa tulla se, että elementtejä on vaihdettu
                    if containsRids == False:
                        rid = dictTranslations[frameText]

            # Koska append tapahtuu vasta seuraavan rid:n löytyessä,
            # täytyy muistaa lisätä myös iteroinnin päätteeksi
            if boolAnnotated:
                lstGoodSearches.append((rid, korp_id))

            #if rid == "78392":
            #    print(rid)
            #    print(korp_ids)
            #    print(boolAnnotated)
            #    print(annotated)
            #    raise Exception
            
            if annotated == 0:
                for korp_id in korp_ids:
                    #if rid == "78392":
                    #    print(rid)
                    #    print(korp_id)
                    #    print("* * * * * * * *  *")
                    lstBadSearches.append((rid, korp_id))
                    
                       
        except Exception as exc:
            print("Virhe tiedostossa: " + filename + ", rivi: " + str(row))
            print("Frame : " + frame)
            print("ENLU : " + ENLU)
            
            #print(frameText)
            raise Exception(str(exc))
            
                                                                        

    return lstGoodSearches, lstBadSearches
